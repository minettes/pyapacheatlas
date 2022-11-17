import itertools
import json
import os
from xml.etree.ElementTree import tostring

from openpyxl import Workbook
from openpyxl import load_workbook

# PyApacheAtlas packages
# Connect to Atlas via a Service Principal
from pyapacheatlas.auth import ServicePrincipalAuthentication
from pyapacheatlas.core import PurviewClient  # Communicate with your Atlas server
from pyapacheatlas.readers import ExcelConfiguration, ExcelReader

# *********
# IMPORTANT
#min standardized
environment = "serving"

if environment == "standardized":
    filter_collection = "ukzobj" #The collection you want to extract
    storage_account ="https://sacontosohealth.dfs.core.windows.net/"

else:
    #min hospital Serving
    filter_collection = "ukzobj" #The collection you want to extract fonterra serving  
    storage_account =  "mssql://fon-dp-dev-srv-syn-003.database.windows.net/"

    # *********

# Modify this filter if you want to add more parameters in the search
custom_filter = {
                    "and": [
                        {"collectionId": filter_collection},
                       # {"entityType": "azure_datalake_gen2_resource_set"},
                        {
                            "not": {
                                "or": [
                                    {
                                        "attributeName": "size",
                                        "operator": "eq",
                                        "attributeValue": 0
                                    },
                                    {
                                        "attributeName": "fileSize",
                                        "operator": "eq",
                                        "attributeValue": 0
                                    },
                                    { "entityType": "azure_datalake_gen2_path"}
                                   
                                ]
                            }
                        }
                    ]
                }

if __name__ == "__main__":
    ""
    
    # Authenticate against your Atlas server
    oauth = ServicePrincipalAuthentication(
        tenant_id=os.environ.get("TENANT_ID", "TENANT_ID"),
        client_id=os.environ.get("CLIENT_ID", "CLIENT_ID"),
        client_secret=os.environ.get("CLIENT_SECRET", "CLIENT_SECRET")
    )
    client = PurviewClient(
        account_name = os.environ.get("PURVIEW_NAME", "PURVIEW_NAME"),
        authentication=oauth
    )

    # When you know the GUID that you want to delete
    # response = client.delete_entity(guid="123-abc-456-def")
    # print(json.dumps(response, indent=2))

    # Replace <YOUR_FQN> with the FQN of the data source assets to be deleted
    # print(custom_filter)
    assets_to_extract = client.search_entities(
        storage_account,
        search_filter=custom_filter
    )

  

    file_path = "./ch_test_download_"+environment+".xlsx"
    excel_config = ExcelConfiguration()
    excel_reader = ExcelReader(excel_config)

    # Create an empty excel template to be populated
    excel_reader.make_template(file_path)


#-----------------------------------------------------
    table_row_counter = 2

    wb = load_workbook(file_path)
    bulkEntity_sheet = wb[excel_config.bulkEntity_sheet]

    for asset in assets_to_extract:
    #Print the whole asset JSON
        

        bulkEntity_sheet.cell(table_row_counter,1).value = asset.get("id")
        bulkEntity_sheet.cell(table_row_counter,2).value = asset.get("entityType")
        bulkEntity_sheet.cell(table_row_counter,3).value = asset.get("name")
        bulkEntity_sheet.cell(table_row_counter,4).value = asset.get("qualifiedName")
        bulkEntity_sheet.cell(table_row_counter,5).value = asset.get("description")

        contacts = asset.get("contact")
        #print(contacts)
        if contacts != None:
            experts = ""
            owners = ""
            for contact in contacts:
                if(contact["contactType"]=="Owner"):
                    owners = owners + contact["id"] +";"
                else:
                    experts = experts + contact["id"] + ";"
     
            bulkEntity_sheet.cell(table_row_counter,6).value = owners
            bulkEntity_sheet.cell(table_row_counter,7).value = experts

   
        classifications = client.get_entity_classifications(asset.get("id"))
        print("----------",asset.get("name"),"--------------")
        #print(json.dumps(classifications, indent=2))

        if classifications != None:
            classifiers = ""
            for classification in (classifications.get("list")):
                if classification.get("source") != "LabelService":
                    classifiers = classifiers + classification.get("typeName") +";"
                    bulkEntity_sheet.cell(table_row_counter,8).value = classifiers

        terms = asset.get("term")
        if terms != None:
            glossaryTerms = ""
            for term in terms:
                glossaryTerms = glossaryTerms + term["name"]+";"
            bulkEntity_sheet.cell(table_row_counter,9).value = glossaryTerms
           
           

        table_row_counter +=1
    wb.save(file_path)

    print("assets_extracted count=" + str(table_row_counter))

results = client.update_businessMetadata()